#!/usr/bin/env python3
"""
Create Excel Upload Template for Workforce Scheduler
Run this script to generate the employee upload template file
Last Updated: 2025-09-09
"""

import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_employee_upload_template():
    """
    Creates an Excel template file for employee uploads with:
    - Base employee columns
    - 10 generic qualification columns
    - Sample data
    - Instructions sheet
    - Data validation
    """
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Employee Data'
    
    # Define headers - YOUR EXACT REQUIREMENTS
    base_headers = [
        'Last Name',
        'First Name', 
        'Employee ID',
        'Crew Assigned',
        'Current Job Position',
        'Email'
    ]
    
    # Generic qualification headers that users can customize
    qualification_headers = [
        'Add Qualification 1',
        'Add Qualification 2',
        'Add Qualification 3',
        'Add Qualification 4',
        'Add Qualification 5',
        'Add Qualification 6',
        'Add Qualification 7',
        'Add Qualification 8',
        'Add Qualification 9',
        'Add Qualification 10'
    ]
    
    # Combine all headers
    all_headers = base_headers + qualification_headers
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers with formatting
    for col, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # Set column widths
        if header == 'Email':
            ws.column_dimensions[get_column_letter(col)].width = 30
        elif header in ['Last Name', 'First Name', 'Current Job Position']:
            ws.column_dimensions[get_column_letter(col)].width = 20
        elif header == 'Employee ID':
            ws.column_dimensions[get_column_letter(col)].width = 15
        elif header == 'Crew Assigned':
            ws.column_dimensions[get_column_letter(col)].width = 12
        else:  # Qualification columns
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Set header row height
    ws.row_dimensions[1].height = 40
    
    # Add crew validation (A, B, C, D)
    crew_validation = DataValidation(
        type="list",
        formula1='"A,B,C,D"',
        allow_blank=False,
        showDropDown=True,
        errorTitle='Invalid Crew',
        error='Please select A, B, C, or D'
    )
    crew_validation.add('D2:D1000')
    ws.add_data_validation(crew_validation)
    
    # Add Yes/No validation for qualification columns
    yes_no_validation = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=True,
        showDropDown=True,
        errorTitle='Invalid Entry',
        error='Please select Yes or No (or leave blank)'
    )
    
    # Apply to qualification columns (columns 7-16)
    for col in range(7, 17):
        col_letter = get_column_letter(col)
        yes_no_validation.add(f'{col_letter}2:{col_letter}1000')
    
    ws.add_data_validation(yes_no_validation)
    
    # Add sample data (4 example employees)
    sample_data = [
        # Employee 1 - Experienced Operator
        ['Smith', 'John', 'EMP001', 'A', 'Senior Operator', 'john.smith@company.com',
         'Yes', 'Yes', 'No', 'Yes', 'No', 'No', 'Yes', 'No', 'No', 'Yes'],
        
        # Employee 2 - Lead with multiple skills
        ['Johnson', 'Sarah', 'EMP002', 'B', 'Lead Operator', 'sarah.johnson@company.com',
         'Yes', 'Yes', 'Yes', 'No', 'Yes', 'No', 'No', 'Yes', 'Yes', 'No'],
        
        # Employee 3 - Maintenance Tech
        ['Williams', 'Michael', 'EMP003', 'C', 'Maintenance Technician', 'mike.williams@company.com',
         'No', 'Yes', 'No', 'No', 'No', 'Yes', 'Yes', 'No', 'Yes', 'No'],
        
        # Employee 4 - Supervisor
        ['Brown', 'Emily', 'EMP004', 'D', 'Shift Supervisor', 'emily.brown@company.com',
         'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'No', 'No', 'No', 'Yes', 'Yes'],
    ]
    
    # Style for sample data
    sample_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    sample_font = Font(italic=True, color="666666")
    
    # Add sample data with formatting
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill = sample_fill
            cell.font = sample_font
            cell.border = thin_border
            
            # Center align Yes/No values
            if col_num > 6:
                cell.alignment = Alignment(horizontal="center")
    
    # Add note row
    note_row = row_num + 2
    note_cell = ws.cell(row=note_row, column=1, 
                       value="DELETE THE SAMPLE ROWS ABOVE BEFORE ADDING YOUR REAL DATA")
    note_cell.font = Font(bold=True, color="FF0000", size=12)
    ws.merge_cells(f'A{note_row}:P{note_row}')
    note_cell.alignment = Alignment(horizontal="center")
    
    # =====================================
    # Create Instructions Sheet
    # =====================================
    ws2 = wb.create_sheet('Instructions')
    ws2.column_dimensions['A'].width = 100
    
    instructions = [
        ('EMPLOYEE UPLOAD TEMPLATE INSTRUCTIONS', 'title'),
        ('', 'blank'),
        ('SYSTEM OVERVIEW:', 'section'),
        ('This template uses a header-based qualification system. The column headers define the skills/qualifications available, and you enter Yes/No for each employee.', 'normal'),
        ('', 'blank'),
        
        ('HOW TO USE THIS TEMPLATE:', 'section'),
        ('1. CUSTOMIZE THE QUALIFICATION HEADERS', 'subsection'),
        ('   • Replace "Add Qualification 1", "Add Qualification 2", etc. with your actual skill names', 'normal'),
        ('   • Examples: "Forklift Certified", "OSHA 10", "Bilingual Spanish", "Lead Experience"', 'normal'),
        ('   • You can add more columns after column P if you need more than 10 qualifications', 'normal'),
        ('   • Delete any unused qualification columns to keep your data clean', 'normal'),
        ('', 'blank'),
        
        ('2. ENTER EMPLOYEE DATA', 'subsection'),
        ('   • Last Name: Employee\'s last name (REQUIRED)', 'normal'),
        ('   • First Name: Employee\'s first name (REQUIRED)', 'normal'),
        ('   • Employee ID: Unique identifier for each employee (REQUIRED)', 'normal'),
        ('   • Crew Assigned: Must be A, B, C, or D (REQUIRED)', 'normal'),
        ('   • Current Job Position: Employee\'s job title/role (REQUIRED)', 'normal'),
        ('   • Email: Employee\'s email address (OPTIONAL but needed for login access)', 'normal'),
        ('', 'blank'),
        
        ('3. MARK QUALIFICATIONS', 'subsection'),
        ('   • For each qualification column, enter:', 'normal'),
        ('     - "Yes" (or Y, 1, True, X) = Employee HAS this qualification', 'normal'),
        ('     - "No" (or N, 0, False) = Employee DOES NOT have this qualification', 'normal'),
        ('     - Leave blank = Employee does not have this qualification', 'normal'),
        ('', 'blank'),
        
        ('EXAMPLE QUALIFICATIONS TO TRACK:', 'section'),
        ('Safety Certifications:', 'subsection'),
        ('   • OSHA 10-Hour', 'normal'),
        ('   • OSHA 30-Hour', 'normal'),
        ('   • First Aid/CPR', 'normal'),
        ('   • Confined Space', 'normal'),
        ('   • Fall Protection', 'normal'),
        ('', 'blank'),
        
        ('Equipment Operations:', 'subsection'),
        ('   • Forklift Operator', 'normal'),
        ('   • Crane Certified', 'normal'),
        ('   • Scissor Lift', 'normal'),
        ('   • Machine Operator', 'normal'),
        ('   • CDL License', 'normal'),
        ('', 'blank'),
        
        ('Technical Skills:', 'subsection'),
        ('   • Welding Certified', 'normal'),
        ('   • Electrical License', 'normal'),
        ('   • Plumbing License', 'normal'),
        ('   • HVAC Certified', 'normal'),
        ('   • Maintenance Tech', 'normal'),
        ('', 'blank'),
        
        ('Leadership/Soft Skills:', 'subsection'),
        ('   • Team Lead Experience', 'normal'),
        ('   • Trainer Certified', 'normal'),
        ('   • Bilingual (Spanish)', 'normal'),
        ('   • Quality Inspector', 'normal'),
        ('   • Safety Committee', 'normal'),
        ('', 'blank'),
        
        ('IMPORTANT NOTES:', 'section'),
        ('• All new employees will have password set to: password123', 'important'),
        ('• Employees must change their password on first login', 'important'),
        ('• Email addresses must be unique across all employees', 'important'),
        ('• Employee IDs must be unique', 'important'),
        ('• Crew assignments affect shift schedules (A/B = Days, C/D = Nights typically)', 'important'),
        ('• Skills/qualifications are used for overtime eligibility and crew assignments', 'important'),
        ('', 'blank'),
        
        ('UPLOAD PROCESS:', 'section'),
        ('1. Complete this template with your employee data', 'normal'),
        ('2. Save the file (keep .xlsx format)', 'normal'),
        ('3. Go to Dashboard → Upload Data', 'normal'),
        ('4. Select this file and click Upload', 'normal'),
        ('5. Review any validation errors', 'normal'),
        ('6. Fix errors and re-upload if needed', 'normal'),
        ('', 'blank'),
        
        ('TROUBLESHOOTING:', 'section'),
        ('• If upload fails, check for:', 'normal'),
        ('  - Missing required fields (Last Name, First Name, Employee ID, Crew, Position)', 'normal'),
        ('  - Invalid crew letters (must be A, B, C, or D)', 'normal'),
        ('  - Duplicate Employee IDs', 'normal'),
        ('  - Invalid email formats', 'normal'),
        ('  - Invalid Yes/No values in qualification columns', 'normal'),
        ('', 'blank'),
        
        ('For questions or issues, contact your system administrator.', 'normal'),
    ]
    
    # Format instructions
    row = 1
    for text, style in instructions:
        cell = ws2.cell(row=row, column=1, value=text)
        
        if style == 'title':
            cell.font = Font(bold=True, size=16, color="366092")
        elif style == 'section':
            cell.font = Font(bold=True, size=12, color="000000")
        elif style == 'subsection':
            cell.font = Font(bold=True, size=11, color="444444")
        elif style == 'important':
            cell.font = Font(bold=True, color="FF0000")
        elif style == 'normal':
            cell.font = Font(size=10)
        
        row += 1
    
    # Save the file
    filename = f'employee_upload_template_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(filename)
    
    print(f"✅ Template created successfully: {filename}")
    print("\nFile contains:")
    print("  • Employee Data sheet with 6 base columns + 10 customizable qualification columns")
    print("  • 4 sample employee rows (delete before using)")
    print("  • Data validation for Crew (A/B/C/D) and Qualifications (Yes/No)")
    print("  • Comprehensive instructions sheet")
    print("\nNext steps:")
    print("1. Open the file and customize the qualification column headers")
    print("2. Delete the sample data rows")
    print("3. Add your employee data")
    print("4. Upload to your system")
    
    return filename

if __name__ == "__main__":
    create_employee_upload_template()
