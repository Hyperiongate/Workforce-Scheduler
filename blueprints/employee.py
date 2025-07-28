from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file
from flask_login import login_required, current_user
from models import db, Schedule, VacationCalendar, Employee, Position, Skill, TimeOffRequest, ShiftSwapRequest, CircadianProfile, CoverageNotification, OvertimeHistory, SleepLog, PositionMessage, PositionMessageRead, MaintenanceIssue, MaintenanceUpdate, ShiftTradePost, ShiftTradeProposal, ShiftTrade
from datetime import date, timedelta, datetime
from sqlalchemy import or_, func, case, and_

# Create the blueprint
employee_bp = Blueprint('employee', __name__)

@employee_bp.route('/view-employees-crews')
@login_required
def view_employees_crews():
    """View all employees and their crew assignments"""
    # Get all employees grouped by crew
    crews = {}
    employees = Employee.query.order_by(Employee.crew, Employee.name).all()
    
    for employee in employees:
        crew_name = employee.crew or 'Unassigned'
        if crew_name not in crews:
            crews[crew_name] = []
        crews[crew_name].append(employee)
    
    # Get positions
    positions = Position.query.all()
    
    # Get skills
    skills = Skill.query.all()
    
    # Calculate statistics
    stats = {
        'total_employees': len(employees),
        'total_crews': len([c for c in crews if c != 'Unassigned']),
        'total_supervisors': len([e for e in employees if e.is_supervisor]),
        'unassigned': len(crews.get('Unassigned', []))
    }
    
    return render_template('view_employees_crews.html',
                         crews=crews,
                         positions=positions,
                         skills=skills,
                         stats=stats)

# REMOVED THE DUPLICATE /overtime-management ROUTE - IT'S NOW IN main.py

@employee_bp.route('/overtime-management/export/excel')
@login_required
def export_overtime_excel():
    """Export overtime data to Excel with current filters and sorting"""
    if not current_user.is_supervisor:
        flash('You must be a supervisor to export data.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Get the same parameters as the main view
        search_term = request.args.get('search', '')
        crew_filter = request.args.get('crew', '')
        position_filter = request.args.get('position', '')
        ot_range_filter = request.args.get('ot_range', '')
        
        # Get sorting parameters
        sort_params = []
        for i in range(1, 5):
            sort_field = request.args.get(f'sort{i}', '')
            sort_dir = request.args.get(f'dir{i}', 'asc')
            if sort_field:
                sort_params.append({'field': sort_field, 'direction': sort_dir})
        
        # Calculate date range
        end_date = date.today()
        start_date = end_date - timedelta(weeks=13)
        
        # Build query (same as main view but without pagination)
        query = Employee.query.filter(Employee.id != current_user.id)
        
        # Apply filters
        if search_term:
            query = query.filter(
                or_(
                    Employee.name.ilike(f'%{search_term}%'),
                    Employee.employee_id.ilike(f'%{search_term}%')
                )
            )
        
        if crew_filter:
            query = query.filter(Employee.crew == crew_filter)
        
        if position_filter:
            query = query.filter(Employee.position_id == int(position_filter))
        
        # Get all employees
        employees = query.all()
        
        # Apply OT range filter if specified
        if ot_range_filter:
            filtered_employees = []
            thirteen_weeks_ago = datetime.now().date() - timedelta(weeks=13)
            
            for emp in employees:
                overtime_total = db.session.query(func.sum(OvertimeHistory.overtime_hours)).filter(
                    OvertimeHistory.employee_id == emp.id,
                    OvertimeHistory.week_start_date >= thirteen_weeks_ago
                ).scalar() or 0.0
                
                if ot_range_filter == '0-50' and 0 <= overtime_total <= 50:
                    filtered_employees.append(emp)
                elif ot_range_filter == '50-100' and 50 < overtime_total <= 100:
                    filtered_employees.append(emp)
                elif ot_range_filter == '100-150' and 100 < overtime_total <= 150:
                    filtered_employees.append(emp)
                elif ot_range_filter == '150-200' and 150 < overtime_total <= 200:
                    filtered_employees.append(emp)
                elif ot_range_filter == '200+' and overtime_total > 200:
                    filtered_employees.append(emp)
            
            employees = filtered_employees
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Overtime Report"
        
        # Headers
        headers = ['Employee ID', 'Name', 'Crew', 'Position', 'Date of Hire', 
                  'Current Week OT', '13-Week Total', 'Weekly Average']
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="11998e", end_color="11998e", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Get overtime data and write rows
        row_num = 2
        thirteen_weeks_ago = datetime.now().date() - timedelta(weeks=13)
        
        for emp in employees:
            # Get overtime data
            overtime_total = db.session.query(func.sum(OvertimeHistory.overtime_hours)).filter(
                OvertimeHistory.employee_id == emp.id,
                OvertimeHistory.week_start_date >= thirteen_weeks_ago
            ).scalar() or 0.0
            
            current_week_start = datetime.now().date() - timedelta(days=datetime.now().weekday())
            current_week_ot = db.session.query(func.sum(OvertimeHistory.overtime_hours)).filter(
                OvertimeHistory.employee_id == emp.id,
                OvertimeHistory.week_start_date == current_week_start
            ).scalar() or 0.0
            
            # Write row
            ws.cell(row=row_num, column=1, value=emp.employee_id or f'EMP{emp.id}')
            ws.cell(row=row_num, column=2, value=emp.name)
            ws.cell(row=row_num, column=3, value=emp.crew or 'Unassigned')
            ws.cell(row=row_num, column=4, value=emp.position.name if emp.position else 'No Position')
            ws.cell(row=row_num, column=5, value=emp.hire_date.strftime('%Y-%m-%d') if emp.hire_date else 'N/A')
            ws.cell(row=row_num, column=6, value=round(current_week_ot, 1))
            ws.cell(row=row_num, column=7, value=round(overtime_total, 1))
            ws.cell(row=row_num, column=8, value=round(overtime_total / 13, 1))
            
            # Apply conditional formatting for high overtime
            if overtime_total > 200:
                for col in range(1, 9):
                    ws.cell(row=row_num, column=col).fill = PatternFill(
                        start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"
                    )
            elif overtime_total > 150:
                for col in range(1, 9):
                    ws.cell(row=row_num, column=col).fill = PatternFill(
                        start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"
                    )
            
            row_num += 1
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to BytesIO object
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'overtime_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error exporting data: {str(e)}', 'danger')
        return redirect(url_for('main.overtime_management'))

@employee_bp.route('/vacation/request', methods=['GET', 'POST'])
@login_required
def vacation_request():
    """Handle vacation/time-off requests"""
    if request.method == 'POST':
        request_type = request.form.get('request_type')
        start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d').date()
        end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d').date()
        reason = request.form.get('reason')
        
        # Validate dates
        if start_date < date.today():
            flash('Start date cannot be in the past', 'danger')
            return redirect(url_for('employee.vacation_request'))
        
        if end_date < start_date:
            flash('End date must be after start date', 'danger')
            return redirect(url_for('employee.vacation_request'))
        
        # Calculate days requested
        days_requested = (end_date - start_date).days + 1
        
        # Check available balance
        if request_type == 'vacation' and days_requested > current_user.vacation_days:
            flash(f'Insufficient vacation days. You have {current_user.vacation_days} days available.', 'danger')
            return redirect(url_for('employee.vacation_request'))
        elif request_type == 'sick' and days_requested > current_user.sick_days:
            flash(f'Insufficient sick days. You have {current_user.sick_days} days available.', 'danger')
            return redirect(url_for('employee.vacation_request'))
        elif request_type == 'personal' and days_requested > current_user.personal_days:
            flash(f'Insufficient personal days. You have {current_user.personal_days} days available.', 'danger')
            return redirect(url_for('employee.vacation_request'))
        
        # Create request
        time_off_request = TimeOffRequest(
            employee_id=current_user.id,
            request_type=request_type,
            start_date=start_date,
            end_date=end_date,
            reason=reason,
            status='pending',
            created_at=datetime.now()
        )
        
        db.session.add(time_off_request)
        db.session.commit()
        
        flash('Time off request submitted successfully!', 'success')
        return redirect(url_for('main.employee_dashboard'))
    
    # GET request - show form
    return render_template('vacation_request.html',
                         vacation_days=current_user.vacation_days,
                         sick_days=current_user.sick_days,
                         personal_days=current_user.personal_days)

@employee_bp.route('/shift-marketplace')
@login_required
def shift_marketplace():
    """Shift trading marketplace"""
    # Get active trade posts
    active_posts = ShiftTradePost.query.filter(
        ShiftTradePost.status == 'active',
        ShiftTradePost.shift_date >= date.today()
    ).order_by(ShiftTradePost.shift_date).all()
    
    # Get user's posts
    my_posts = ShiftTradePost.query.filter(
        ShiftTradePost.employee_id == current_user.id,
        ShiftTradePost.shift_date >= date.today()
    ).order_by(ShiftTradePost.shift_date).all()
    
    # Get proposals for user's posts
    my_proposals = ShiftTradeProposal.query.join(ShiftTradePost).filter(
        ShiftTradePost.employee_id == current_user.id,
        ShiftTradeProposal.status == 'pending'
    ).all()
    
    # Get user's pending proposals
    pending_proposals = ShiftTradeProposal.query.filter(
        ShiftTradeProposal.proposer_id == current_user.id,
        ShiftTradeProposal.status == 'pending'
    ).all()
    
    return render_template('shift_marketplace.html',
                         active_posts=active_posts,
                         my_posts=my_posts,
                         my_proposals=my_proposals,
                         pending_proposals=pending_proposals)

@employee_bp.route('/sleep-dashboard')
@login_required
def sleep_dashboard():
    """Sleep health dashboard"""
    # Get or create circadian profile
    profile = CircadianProfile.query.filter_by(employee_id=current_user.id).first()
    
    if not profile:
        # Redirect to profile creation
        return redirect(url_for('employee.sleep_profile'))
    
    # Get recent sleep logs
    sleep_logs = SleepLog.query.filter_by(
        employee_id=current_user.id
    ).order_by(SleepLog.date.desc()).limit(7).all()
    
    # Calculate sleep metrics
    if sleep_logs:
        avg_duration = sum(log.sleep_duration or 0 for log in sleep_logs) / len(sleep_logs)
        avg_quality = sum(log.sleep_quality or 0 for log in sleep_logs) / len(sleep_logs)
    else:
        avg_duration = 0
        avg_quality = 0
    
    return render_template('sleep_dashboard.html',
                         profile=profile,
                         sleep_logs=sleep_logs,
                         avg_duration=avg_duration,
                         avg_quality=avg_quality)

@employee_bp.route('/sleep-profile', methods=['GET', 'POST'])
@login_required
def sleep_profile():
    """Create or update circadian profile"""
    profile = CircadianProfile.query.filter_by(employee_id=current_user.id).first()
    
    if request.method == 'POST':
        chronotype = request.form.get('chronotype')
        preferred_shift = request.form.get('preferred_shift')
        sleep_goal = int(request.form.get('sleep_goal', 8))
        
        if profile:
            profile.chronotype = chronotype
            profile.preferred_shift_type = preferred_shift
            profile.sleep_goal = sleep_goal
            profile.updated_at = datetime.now()
        else:
            profile = CircadianProfile(
                employee_id=current_user.id,
                chronotype=chronotype,
                preferred_shift_type=preferred_shift,
                sleep_goal=sleep_goal,
                assessment_completed=datetime.now()
            )
            db.session.add(profile)
        
        db.session.commit()
        flash('Sleep profile updated successfully!', 'success')
        return redirect(url_for('employee.sleep_dashboard'))
    
    return render_template('sleep_profile_form.html', profile=profile)

@employee_bp.route('/position/messages')
@login_required
def position_messages():
    """View messages for your position"""
    if not current_user.position_id:
        flash('You must have a position assigned to view position messages.', 'warning')
        return redirect(url_for('main.employee_dashboard'))
    
    # Get messages for user's position
    messages = PositionMessage.query.filter_by(
        position_id=current_user.position_id
    ).order_by(PositionMessage.sent_at.desc()).all()
    
    # Mark messages as read
    for message in messages:
        read_receipt = PositionMessageRead.query.filter_by(
            message_id=message.id,
            reader_id=current_user.id
        ).first()
        
        if not read_receipt:
            read_receipt = PositionMessageRead(
                message_id=message.id,
                reader_id=current_user.id,
                read_at=datetime.now()
            )
            db.session.add(read_receipt)
    
    db.session.commit()
    
    return render_template('position_messages.html', messages=messages)

@employee_bp.route('/maintenance/report', methods=['GET', 'POST'])
@login_required
def report_maintenance():
    """Report a maintenance issue"""
    if request.method == 'POST':
        title = request.form.get('title')
        description = request.form.get('description')
        location = request.form.get('location')
        category = request.form.get('category', 'equipment')
        priority = request.form.get('priority', 'medium')
        
        issue = MaintenanceIssue(
            reporter_id=current_user.id,
            title=title,
            description=description,
            location=location,
            category=category,
            priority=priority,
            status='new',
            created_at=datetime.now()
        )
        
        db.session.add(issue)
        db.session.commit()
        
        flash('Maintenance issue reported successfully!', 'success')
        return redirect(url_for('employee.maintenance_issues'))
    
    return render_template('report_maintenance.html')

@employee_bp.route('/maintenance/issues')
@login_required
def maintenance_issues():
    """View maintenance issues"""
    # Get all open issues
    open_issues = MaintenanceIssue.query.filter(
        MaintenanceIssue.status.in_(['new', 'in_progress'])
    ).order_by(
        case(
            (MaintenanceIssue.priority == 'critical', 1),
            (MaintenanceIssue.priority == 'high', 2),
            (MaintenanceIssue.priority == 'medium', 3),
            (MaintenanceIssue.priority == 'low', 4)
        )
    ).all()
    
    # Get user's reported issues
    my_issues = MaintenanceIssue.query.filter_by(
        reporter_id=current_user.id
    ).order_by(MaintenanceIssue.created_at.desc()).limit(10).all()
    
    return render_template('maintenance_issues.html',
                         open_issues=open_issues,
                         my_issues=my_issues)

@employee_bp.route('/swap-request', methods=['POST'])
@login_required
def create_swap_request():
    """Create a shift swap request"""
    requester_shift_id = request.form.get('requester_shift_id')
    target_shift_id = request.form.get('target_shift_id')
    reason = request.form.get('reason')
    
    # Get the shifts
    requester_shift = Schedule.query.get(requester_shift_id)
    target_shift = Schedule.query.get(target_shift_id)
    
    if not requester_shift or not target_shift:
        flash('Invalid shift selection', 'danger')
        return redirect(url_for('employee.shift_marketplace'))
    
    # Verify requester owns the shift
    if requester_shift.employee_id != current_user.id:
        flash('You can only swap your own shifts', 'danger')
        return redirect(url_for('employee.shift_marketplace'))
    
    # Create swap request
    swap_request = ShiftSwapRequest(
        requester_id=current_user.id,
        requested_with_id=target_shift.employee_id,
        requester_shift_date=requester_shift.date,
        requested_shift_date=target_shift.date,
        reason=reason,
        status='pending',
        created_at=datetime.now()
    )
    
    db.session.add(swap_request)
    db.session.commit()
    
    flash('Shift swap request submitted successfully!', 'success')
    return redirect(url_for('employee.shift_marketplace'))

@employee_bp.route('/employees/crew-management')
@login_required
def crew_management():
    """Crew management page - for supervisors"""
    if not current_user.is_supervisor:
        flash('You must be a supervisor to access this page.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Get all employees grouped by crew
    crews = {'A': [], 'B': [], 'C': [], 'D': [], 'Unassigned': []}
    employees = Employee.query.order_by(Employee.crew, Employee.name).all()
    
    for employee in employees:
        crew = employee.crew if employee.crew in ['A', 'B', 'C', 'D'] else 'Unassigned'
        crews[crew].append(employee)
    
    # Get statistics
    stats = {
        'total_employees': len(employees),
        'crew_a': len(crews['A']),
        'crew_b': len(crews['B']),
        'crew_c': len(crews['C']),
        'crew_d': len(crews['D']),
        'unassigned': len(crews['Unassigned'])
    }
    
    return render_template('crew_management.html', crews=crews, stats=stats)

@employee_bp.route('/shift-trade/post', methods=['POST'])
@login_required
def post_shift_trade():
    """Post a shift for trade"""
    shift_date = datetime.strptime(request.form.get('shift_date'), '%Y-%m-%d').date()
    shift_type = request.form.get('shift_type')
    reason = request.form.get('reason')
    
    # Validate future date
    if shift_date <= date.today():
        flash('You can only post future shifts for trade', 'danger')
        return redirect(url_for('employee.shift_marketplace'))
    
    # Create trade post
    post = ShiftTradePost(
        employee_id=current_user.id,
        shift_date=shift_date,
        shift_type=shift_type,
        reason=reason,
        status='active',
        posted_at=datetime.now()
    )
    
    db.session.add(post)
    db.session.commit()
    
    flash('Shift posted for trade successfully!', 'success')
    return redirect(url_for('employee.shift_marketplace'))

@employee_bp.route('/shift-trade/propose/<int:post_id>', methods=['POST'])
@login_required
def propose_trade(post_id):
    """Propose a trade for a posted shift"""
    post = ShiftTradePost.query.get_or_404(post_id)
    proposed_date = datetime.strptime(request.form.get('proposed_date'), '%Y-%m-%d').date()
    message = request.form.get('message')
    
    # Validate
    if post.employee_id == current_user.id:
        flash('You cannot propose a trade for your own shift', 'danger')
        return redirect(url_for('employee.shift_marketplace'))
    
    if proposed_date <= date.today():
        flash('Proposed date must be in the future', 'danger')
        return redirect(url_for('employee.shift_marketplace'))
    
    # Create proposal
    proposal = ShiftTradeProposal(
        post_id=post_id,
        proposer_id=current_user.id,
        proposed_date=proposed_date,
        message=message,
        status='pending',
        proposed_at=datetime.now()
    )
    
    db.session.add(proposal)
    db.session.commit()
    
    flash('Trade proposal submitted successfully!', 'success')
    return redirect(url_for('employee.shift_marketplace'))

@employee_bp.route('/shift-trade/accept/<int:proposal_id>', methods=['POST'])
@login_required
def accept_trade_proposal(proposal_id):
    """Accept a trade proposal"""
    proposal = ShiftTradeProposal.query.get_or_404(proposal_id)
    post = proposal.post
    
    # Validate ownership
    if post.employee_id != current_user.id:
        flash('You can only accept proposals for your own posts', 'danger')
        return redirect(url_for('employee.shift_marketplace'))
    
    # Create the trade
    trade = ShiftTrade(
        employee1_id=post.employee_id,
        employee2_id=proposal.proposer_id,
        shift1_date=post.shift_date,
        shift2_date=proposal.proposed_date,
        status='pending_approval',
        created_at=datetime.now()
    )
    
    db.session.add(trade)
    
    # Update proposal and post status
    proposal.status = 'accepted'
    post.status = 'matched'
    
    # Reject other proposals for this post
    other_proposals = ShiftTradeProposal.query.filter(
        ShiftTradeProposal.post_id == post_id,
        ShiftTradeProposal.id != proposal_id,
        ShiftTradeProposal.status == 'pending'
    ).all()
    
    for other in other_proposals:
        other.status = 'rejected'
    
    db.session.commit()
    
    flash('Trade proposal accepted! Awaiting supervisor approval.', 'success')
    return redirect(url_for('employee.shift_marketplace'))

# API endpoints for AJAX functionality
@employee_bp.route('/api/overtime-week-details/<int:employee_id>')
@login_required
def overtime_week_details(employee_id):
    """Get detailed overtime data for an employee"""
    if not current_user.is_supervisor and current_user.id != employee_id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    # Get 13 weeks of data
    weeks = []
    for i in range(13):
        week_start = date.today() - timedelta(weeks=i, days=date.today().weekday())
        week_end = week_start + timedelta(days=6)
        
        overtime = OvertimeHistory.query.filter_by(
            employee_id=employee_id,
            week_start_date=week_start
        ).first()
        
        weeks.append({
            'week': f'Week of {week_start.strftime("%b %d")}',
            'hours': overtime.overtime_hours if overtime else 0
        })
    
    return jsonify({'weeks': list(reversed(weeks))})

@employee_bp.route('/api/position-colleagues')
@login_required
def get_position_colleagues():
    """Get colleagues in same position but different crews"""
    if not current_user.position_id:
        return jsonify([])
    
    colleagues = Employee.query.filter(
        Employee.position_id == current_user.position_id,
        Employee.id != current_user.id,
        Employee.crew != current_user.crew
    ).all()
    
    return jsonify([{
        'id': c.id,
        'name': c.name,
        'crew': c.crew
    } for c in colleagues])
