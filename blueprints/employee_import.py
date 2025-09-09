# blueprints/employee_import.py
"""
Complete Excel upload system for employee data management
FIXED VERSION - Deploy this ENTIRE file to blueprints/employee_import.py
Last Updated: 2025-09-08
Changes: Added all missing routes for validation, export, download, and AJAX endpoints
"""

from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, current_app, send_file, make_response, abort
from flask_login import login_required, current_user
from models import db, Employee, Position, OvertimeHistory, FileUpload, Skill, EmployeeSkill
from datetime import datetime, timedelta, date
from werkzeug.utils import secure_filename
from sqlalchemy import func, and_, or_, text
from functools import wraps
import pandas as pd
import os
import io
import re
import logging
import json
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Set up logging
logger = logging.getLogger(__name__)

# Create blueprint - NO url_prefix so routes are at root level
employee_import_bp = Blueprint('employee_import', __name__)

# ==========================================
# DECORATORS AND HELPERS
# ==========================================

def supervisor_required(f):
    """Decorator to require supervisor access"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('auth.login'))
        if not current_user.is_supervisor:
            flash('Access denied. Supervisors only.', 'error')
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def allowed_file(filename):
    """Check if uploaded file has allowed extension"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls'}

def secure_file_path(filename):
    """Generate secure file path for uploads"""
    upload_folder = current_app.config.get('UPLOAD_FOLDER', 'upload_files')
    if not os.path.isabs(upload_folder):
        upload_folder = os.path.join(current_app.root_path, upload_folder)
    
    os.makedirs(upload_folder, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    name, ext = os.path.splitext(secure_filename(filename))
    unique_filename = f"{name}_{timestamp}{ext}"
    
    return os.path.join(upload_folder, unique_filename)

# ==========================================
# HELPER FUNCTIONS FOR STATS AND DATA
# ==========================================

def get_employee_stats():
    """Get employee statistics for dashboard"""
    try:
        total_employees = Employee.query.filter_by(is_active=True).count()
        
        crews = db.session.query(
            Employee.crew, 
            func.count(Employee.id)
        ).filter(
            Employee.is_active == True
        ).group_by(Employee.crew).all()
        
        crew_distribution = {crew: count for crew, count in crews}
        
        # Get skill statistics
        total_skills = Skill.query.count()
        employees_with_skills = db.session.query(
            func.count(func.distinct(EmployeeSkill.employee_id))
        ).scalar() or 0
        
        # Get overtime statistics
        employees_with_ot = db.session.query(
            func.count(func.distinct(OvertimeHistory.employee_id))
        ).scalar() or 0
        
        # Calculate OT categories
        low_ot = 0
        medium_ot = 0
        high_ot = 0
        
        # Get average OT per employee
        ot_data = db.session.query(
            OvertimeHistory.employee_id,
            func.avg(OvertimeHistory.overtime_hours).label('avg_ot')
        ).group_by(OvertimeHistory.employee_id).all()
        
        for emp_ot in ot_data:
            if emp_ot.avg_ot < 10:
                low_ot += 1
            elif emp_ot.avg_ot < 20:
                medium_ot += 1
            else:
                high_ot += 1
        
        return {
            'total_employees': total_employees,
            'crews': crew_distribution,
            'total_skills': total_skills,
            'employees_with_skills': employees_with_skills,
            'with_overtime': employees_with_ot,
            'low_ot': low_ot,
            'medium_ot': medium_ot,
            'high_ot': high_ot,
            'last_updated': datetime.now()
        }
    except Exception as e:
        logger.error(f"Error getting employee stats: {e}")
        return {
            'total_employees': 0,
            'crews': {},
            'total_skills': 0,
            'employees_with_skills': 0,
            'with_overtime': 0,
            'low_ot': 0,
            'medium_ot': 0,
            'high_ot': 0,
            'last_updated': datetime.now()
        }

def get_recent_uploads(limit=5):
    """Get recent upload history"""
    try:
        uploads = FileUpload.query.order_by(
            FileUpload.uploaded_at.desc()
        ).limit(limit).all()
        
        return [{
            'id': upload.id,
            'filename': upload.filename,
            'upload_type': upload.upload_type or 'employee',
            'status': upload.status or 'completed',
            'records_processed': upload.records_processed or 0,
            'created_at': upload.uploaded_at,
            'uploaded_by': upload.uploaded_by.name if upload.uploaded_by else 'Unknown'
        } for upload in uploads]
    except Exception as e:
        logger.error(f"Error getting recent uploads: {e}")
        return []

def create_or_get_skill(skill_name, category='General'):
    """Create a skill if it doesn't exist, or return existing one"""
    skill_name = skill_name.strip()
    skill = Skill.query.filter_by(name=skill_name).first()
    
    if not skill:
        # Determine category based on skill name
        if any(cert in skill_name.upper() for cert in ['OSHA', 'CPR', 'FIRST AID', 'CERTIFIED', 'LICENSE']):
            category = 'Certification'
        elif any(equip in skill_name.lower() for equip in ['forklift', 'crane', 'machine', 'equipment']):
            category = 'Equipment'
        elif any(tech in skill_name.lower() for tech in ['welding', 'electrical', 'plumbing', 'hvac']):
            category = 'Technical'
        elif any(mgmt in skill_name.lower() for mgmt in ['leadership', 'supervisor', 'management', 'six sigma', 'pmp']):
            category = 'Management'
        
        skill = Skill(
            name=skill_name,
            category=category,
            requires_renewal='certified' in skill_name.lower() or 'license' in skill_name.lower()
        )
        db.session.add(skill)
        db.session.flush()
    
    return skill

def assign_skill_to_employee(employee, skill, certified_date=None):
    """Assign a skill to an employee if not already assigned"""
    existing = EmployeeSkill.query.filter_by(
        employee_id=employee.id,
        skill_id=skill.id
    ).first()
    
    if not existing:
        employee_skill = EmployeeSkill(
            employee_id=employee.id,
            skill_id=skill.id,
            certified_date=certified_date or date.today()
        )
        
        # Set expiry for renewable certifications (1 year default)
        if skill.requires_renewal:
            employee_skill.expiry_date = (certified_date or date.today()) + timedelta(days=365)
        
        db.session.add(employee_skill)
        return True
    return False

# ==========================================
# VALIDATION FUNCTIONS WITH HEADER-BASED QUALIFICATIONS
# ==========================================

def validate_employee_data_comprehensive(df):
    """Validate employee data - headers define qualifications, cells contain Yes/No"""
    errors = []
    warnings = []
    
    # YOUR required columns
    your_required_columns = ['Last Name', 'First Name', 'Employee ID', 'Crew Assigned', 'Current Job Position']
    your_optional_columns = ['Email']
    
    # Identify qualification columns (any column after the base columns)
    base_columns = your_required_columns + your_optional_columns
    qualification_columns = []
    
    for col in df.columns:
        if col not in base_columns and col.strip() != '':
            qualification_columns.append(col)
    
    logger.info(f"Found {len(qualification_columns)} qualification columns: {qualification_columns}")
    
    # Check required columns exist
    missing_columns = []
    for col in your_required_columns:
        if col not in df.columns:
            missing_columns.append(col)
    
    if missing_columns:
        return {
            'success': False,
            'error': f'Missing required columns: {", ".join(missing_columns)}',
            'errors': [f'Missing column: {col}' for col in missing_columns]
        }
    
    # Validate each row
    seen_employee_ids = set()
    valid_crews = {'A', 'B', 'C', 'D'}
    
    for idx, row in df.iterrows():
        row_num = idx + 2  # Excel row number (header is row 1)
        
        # Check Last Name
        if pd.isna(row.get('Last Name')) or str(row.get('Last Name')).strip() == '':
            errors.append(f'Row {row_num}: Missing Last Name')
        
        # Check First Name
        if pd.isna(row.get('First Name')) or str(row.get('First Name')).strip() == '':
            errors.append(f'Row {row_num}: Missing First Name')
        
        # Check Employee ID
        emp_id = row.get('Employee ID')
        if pd.isna(emp_id) or str(emp_id).strip() == '':
            errors.append(f'Row {row_num}: Missing Employee ID')
        else:
            emp_id_str = str(emp_id).strip()
            if emp_id_str in seen_employee_ids:
                errors.append(f'Row {row_num}: Duplicate Employee ID "{emp_id_str}"')
            seen_employee_ids.add(emp_id_str)
        
        # Check Crew
        crew = row.get('Crew Assigned')
        if pd.isna(crew) or str(crew).strip() == '':
            errors.append(f'Row {row_num}: Missing Crew Assigned')
        elif str(crew).strip().upper() not in valid_crews:
            errors.append(f'Row {row_num}: Invalid Crew "{crew}" (must be A, B, C, or D)')
        
        # Check Position
        if pd.isna(row.get('Current Job Position')) or str(row.get('Current Job Position')).strip() == '':
            errors.append(f'Row {row_num}: Missing Current Job Position')
        
        # Check Email (optional but validate format if provided)
        email = row.get('Email')
        if not pd.isna(email) and str(email).strip() != '':
            email_str = str(email).strip()
            if '@' not in email_str:
                warnings.append(f'Row {row_num}: Invalid email format "{email_str}"')
        
        # Validate qualification values (should be Yes, No, Y, N, or blank)
        valid_qual_values = {'yes', 'y', 'no', 'n', '1', '0', 'true', 'false', 'x', ''}
        for qual_col in qualification_columns:
            qual_value = row.get(qual_col)
            if not pd.isna(qual_value) and str(qual_value).strip() != '':
                val_lower = str(qual_value).strip().lower()
                if val_lower not in valid_qual_values:
                    warnings.append(f'Row {row_num}, {qual_col}: Invalid value "{qual_value}" (use Yes/No or leave blank)')
    
    # Check if dataframe is empty
    if len(df) == 0:
        errors.append('No data rows found in file')
    
    # Return validation results
    if errors:
        return {
            'success': False,
            'error': f'Validation failed with {len(errors)} errors',
            'errors': errors[:20],  # Limit to first 20 errors
            'total_errors': len(errors),
            'warnings': warnings
        }
    
    # Count employees with qualifications
    employees_with_quals = 0
    for idx, row in df.iterrows():
        has_qual = False
        for qual_col in qualification_columns:
            val = row.get(qual_col)
            if not pd.isna(val) and str(val).strip().lower() in ['yes', 'y', '1', 'true', 'x']:
                has_qual = True
                break
        if has_qual:
            employees_with_quals += 1
    
    return {
        'success': True,
        'message': 'Validation passed',
        'employee_count': len(df),
        'warnings': warnings,
        'qualification_columns': qualification_columns,
        'qualification_count': len(qualification_columns),
        'employees_with_qualifications': employees_with_quals
    }

# ==========================================
# VALIDATION ROUTE - ADDED 2025-01-09
# ==========================================

@employee_import_bp.route('/validate-upload', methods=['POST'])
@login_required
@supervisor_required
def validate_upload():
    """AJAX endpoint to validate uploaded file without importing"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'})
        
        file = request.files['file']
        upload_type = request.form.get('uploadType', 'employee')
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'})
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'error': 'Invalid file type. Please upload .xlsx or .xls'})
        
        # Save file temporarily
        filepath = secure_file_path(file.filename)
        file.save(filepath)
        
        try:
            # Read Excel file
            df = pd.read_excel(filepath, sheet_name=None)
            
            # Get the appropriate sheet
            if isinstance(df, dict):
                if 'Employee Data' in df:
                    df = df['Employee Data']
                elif 'Overtime Data' in df:
                    df = df['Overtime Data']
                else:
                    df = df[list(df.keys())[0]]
            
            # Validate based on type
            if upload_type == 'employee':
                result = validate_employee_data_comprehensive(df)
            elif upload_type == 'overtime':
                result = validate_overtime_data(df)
            else:
                result = {'success': False, 'error': 'Invalid upload type'}
            
            # Add row count
            result['total_rows'] = len(df)
            
            # Clean up temp file
            if os.path.exists(filepath):
                os.remove(filepath)
            
            return jsonify(result)
            
        except Exception as e:
            logger.error(f"Error validating file: {e}")
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({'success': False, 'error': f'Error reading file: {str(e)}'})
            
    except Exception as e:
        logger.error(f"Error in validate_upload: {e}")
        return jsonify({'success': False, 'error': 'Server error during validation'})

def validate_overtime_data(df):
    """Validate overtime data"""
    errors = []
    warnings = []
    
    # Required columns for overtime
    required_cols = ['Employee ID', 'Week Start Date', 'Regular Hours', 'Overtime Hours']
    
    # Check for missing columns
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        return {
            'success': False,
            'error': f'Missing required columns: {", ".join(missing_cols)}',
            'errors': [f'Missing column: {col}' for col in missing_cols]
        }
    
    # Track employees and their weeks
    employee_weeks = {}
    total_ot_hours = 0
    
    for idx, row in df.iterrows():
        row_num = idx + 2
        
        # Validate Employee ID
        emp_id = row.get('Employee ID')
        if pd.isna(emp_id) or str(emp_id).strip() == '':
            errors.append(f'Row {row_num}: Missing Employee ID')
            continue
        
        emp_id_str = str(emp_id).strip()
        
        # Check if employee exists in database
        employee = Employee.query.filter_by(employee_id=emp_id_str).first()
        if not employee:
            errors.append(f'Row {row_num}: Employee ID "{emp_id_str}" not found in system')
        
        # Validate Week Start Date
        try:
            week_date = pd.to_datetime(row['Week Start Date'])
            if week_date.weekday() != 0:  # Monday is 0
                warnings.append(f'Row {row_num}: Week start date is not a Monday')
        except:
            errors.append(f'Row {row_num}: Invalid date format for Week Start Date')
            continue
        
        # Track weeks per employee
        if emp_id_str not in employee_weeks:
            employee_weeks[emp_id_str] = []
        employee_weeks[emp_id_str].append(week_date)
        
        # Validate hours
        try:
            regular_hours = float(row.get('Regular Hours', 0))
            overtime_hours = float(row.get('Overtime Hours', 0))
            
            if regular_hours < 0:
                errors.append(f'Row {row_num}: Regular hours cannot be negative')
            if overtime_hours < 0:
                errors.append(f'Row {row_num}: Overtime hours cannot be negative')
            
            # Warnings for unusual values
            if regular_hours > 60:
                warnings.append(f'Row {row_num}: Unusually high regular hours ({regular_hours})')
            if overtime_hours > 40:
                warnings.append(f'Row {row_num}: Very high overtime hours ({overtime_hours})')
            
            total_ot_hours += overtime_hours
            
        except (ValueError, TypeError):
            errors.append(f'Row {row_num}: Invalid numeric value for hours')
    
    # Check for 13 weeks per employee
    for emp_id, weeks in employee_weeks.items():
        if len(weeks) != 13:
            warnings.append(f'Employee {emp_id} has {len(weeks)} weeks (expected 13)')
    
    # Calculate average OT
    avg_ot = total_ot_hours / len(df) if len(df) > 0 else 0
    
    if errors:
        return {
            'success': False,
            'error': f'Validation failed with {len(errors)} errors',
            'errors': errors[:20],
            'total_errors': len(errors),
            'warnings': warnings
        }
    
    return {
        'success': True,
        'message': 'Overtime validation passed',
        'employee_count': len(employee_weeks),
        'total_rows': len(df),
        'avg_ot': round(avg_ot, 1),
        'warnings': warnings
    }

# ==========================================
# PROCESSING FUNCTIONS WITH HEADER-BASED SKILLS
# ==========================================

def process_employee_upload(df, upload_record, replace_all=False):
    """Process employee data - column headers are skill names, cells are Yes/No"""
    try:
        created_count = 0
        updated_count = 0
        skipped_count = 0
        skills_added = 0
        
        # Identify qualification columns
        base_columns = ['Last Name', 'First Name', 'Employee ID', 'Crew Assigned', 'Current Job Position', 'Email']
        qualification_columns = [col for col in df.columns if col not in base_columns and col.strip() != '']
        
        logger.info(f"Processing with qualification columns: {qualification_columns}")
        
        # Create skills for each qualification column
        skill_map = {}
        for qual_col in qualification_columns:
            skill = create_or_get_skill(qual_col)
            skill_map[qual_col] = skill
            logger.info(f"Created/found skill: {qual_col}")
        
        # If replace_all, deactivate all existing employees first
        if replace_all:
            Employee.query.update({'is_active': False})
            logger.info("Deactivated all existing employees for replacement")
        
        # Process each row
        for idx, row in df.iterrows():
            try:
                # Extract data with YOUR column names
                emp_id = str(row['Employee ID']).strip()
                first_name = str(row['First Name']).strip()
                last_name = str(row['Last Name']).strip()
                full_name = f"{first_name} {last_name}"
                crew = str(row['Crew Assigned']).strip().upper()
                position_name = str(row['Current Job Position']).strip()
                email = str(row.get('Email', '')).strip() if not pd.isna(row.get('Email')) else None
                
                # Check if employee exists
                employee = Employee.query.filter_by(employee_id=emp_id).first()
                
                if employee:
                    # Update existing employee
                    employee.name = full_name
                    employee.crew = crew
                    employee.is_active = True
                    if email:
                        employee.email = email
                    updated_count += 1
                    logger.info(f"Updated employee: {emp_id}")
                else:
                    # Create new employee with default password
                    employee = Employee(
                        employee_id=emp_id,
                        name=full_name,
                        crew=crew,
                        email=email or f"{emp_id}@company.local",  # Default email if none provided
                        is_active=True
                    )
                    employee.set_password('password123')  # Default password
                    db.session.add(employee)
                    created_count += 1
                    logger.info(f"Created new employee: {emp_id}")
                
                # Handle position
                position = Position.query.filter_by(name=position_name).first()
                if not position:
                    position = Position(name=position_name, department='Operations')
                    db.session.add(position)
                    logger.info(f"Created new position: {position_name}")
                
                employee.position = position
                
                # Process qualifications based on Yes/No values
                for qual_col, skill in skill_map.items():
                    qual_value = row.get(qual_col)
                    if not pd.isna(qual_value) and str(qual_value).strip() != '':
                        val_lower = str(qual_value).strip().lower()
                        # Check if this is a "Yes" value
                        if val_lower in ['yes', 'y', '1', 'true', 'x']:
                            if assign_skill_to_employee(employee, skill):
                                skills_added += 1
                                logger.info(f"Added skill '{qual_col}' to employee {emp_id}")
                
                # Commit after each employee to avoid losing all on error
                db.session.commit()
                
            except Exception as e:
                logger.error(f"Error processing row {idx + 2}: {e}")
                db.session.rollback()
                skipped_count += 1
                continue
        
        # Final commit
        db.session.commit()
        
        return {
            'success': True,
            'created': created_count,
            'updated': updated_count,
            'skipped': skipped_count,
            'skills_added': skills_added,
            'records_processed': len(df),
            'message': f'Successfully processed {len(df)} records. Created: {created_count}, Updated: {updated_count}, Skills: {skills_added}'
        }
        
    except Exception as e:
        logger.error(f"Error in process_employee_upload: {e}")
        db.session.rollback()
        return {
            'success': False,
            'error': str(e),
            'records_processed': 0
        }

# ==========================================
# MAIN ROUTES
# ==========================================

@employee_import_bp.route('/upload-employees')
@login_required
@supervisor_required
def upload_employees():
    """Main upload page for employees"""
    try:
        stats = get_employee_stats()
        recent_uploads = get_recent_uploads()
        
        # Try to use your simple template first
        template_options = [
            'upload_employees_simple.html',
            'upload_employees_simple_direct.html',
            'upload_employees_enhanced.html',
            'upload_employees.html',
            'employee_upload.html'
        ]
        
        for template_name in template_options:
            try:
                return render_template(
                    template_name,
                    stats=stats,
                    recent_uploads=recent_uploads,
                    crew_distribution=stats.get('crews', {}),
                    total_employees=stats.get('total_employees', 0),
                    total_skills=stats.get('total_skills', 0),
                    employees_with_skills=stats.get('employees_with_skills', 0)
                )
            except Exception:
                continue
        
        # If no template works, render inline HTML
        return render_simple_upload_page(stats, recent_uploads)
        
    except Exception as e:
        logger.error(f"Error in upload_employees: {e}")
        flash('Error loading upload page. Please try again.', 'error')
        return redirect(url_for('supervisor.dashboard'))

@employee_import_bp.route('/upload-overtime')
@login_required
@supervisor_required
def upload_overtime():
    """Upload overtime data page"""
    try:
        stats = get_employee_stats()
        recent_uploads = get_recent_uploads()
        
        # Try to use overtime template if it exists
        try:
            return render_template(
                'upload_overtime.html',
                stats=stats,
                recent_uploads=recent_uploads
            )
        except:
            # Fall back to employee upload page with a message
            flash('Overtime upload functionality coming soon. Use the template below for now.', 'info')
            return redirect(url_for('employee_import.upload_employees'))
            
    except Exception as e:
        logger.error(f"Error in upload_overtime: {e}")
        flash('Error loading overtime upload page.', 'error')
        return redirect(url_for('supervisor.dashboard'))

@employee_import_bp.route('/upload-employees', methods=['POST'])
@login_required
@supervisor_required
def upload_employees_post():
    """Process employee upload"""
    try:
        if 'file' not in request.files:
            flash('No file provided', 'error')
            return redirect(url_for('employee_import.upload_employees'))
        
        file = request.files['file']
        upload_type = request.form.get('uploadType', 'employee')
        replace_all = 'replaceAll' in request.form
        
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('employee_import.upload_employees'))
        
        if not allowed_file(file.filename):
            flash('Please upload an Excel file (.xlsx or .xls)', 'error')
            return redirect(url_for('employee_import.upload_employees'))
        
        # Save file
        filepath = secure_file_path(file.filename)
        file.save(filepath)
        
        # Create upload record
        upload_record = FileUpload(
            filename=file.filename,
            upload_type=upload_type,
            uploaded_by_id=current_user.id,
            status='processing',
            file_size=os.path.getsize(filepath),
            file_path=filepath
        )
        db.session.add(upload_record)
        db.session.flush()
        
        try:
            # Read Excel file
            df = pd.read_excel(filepath, sheet_name=None)
            
            # Get the appropriate sheet
            if isinstance(df, dict):
                if 'Employee Data' in df:
                    df = df['Employee Data']
                elif 'Overtime Data' in df:
                    df = df['Overtime Data']
                else:
                    df = df[list(df.keys())[0]]
            
            # Validate data
            if upload_type == 'employee':
                validation = validate_employee_data_comprehensive(df)
                if validation['success']:
                    result = process_employee_upload(df, upload_record, replace_all)
                else:
                    result = validation
            else:
                result = {'success': False, 'error': 'Invalid upload type'}
            
            # Update upload record
            if result.get('success'):
                upload_record.status = 'completed'
                upload_record.records_processed = result.get('records_processed', 0)
                upload_record.successful_records = result.get('created', 0) + result.get('updated', 0)
                
                # Build success message
                msg_parts = [f"Successfully processed {result.get('records_processed', 0)} records!"]
                if result.get('created'):
                    msg_parts.append(f"{result.get('created')} created")
                if result.get('updated'):
                    msg_parts.append(f"{result.get('updated')} updated")
                if result.get('skills_added'):
                    msg_parts.append(f"{result.get('skills_added')} skills assigned")
                
                flash(' - '.join(msg_parts), 'success')
            else:
                upload_record.status = 'failed'
                upload_record.error_details = json.dumps({'error': result.get('error'), 'errors': result.get('errors', [])})
                flash(f"Upload failed: {result.get('error')}", 'error')
            
            db.session.commit()
            
        except Exception as e:
            logger.error(f"Error processing upload: {e}\n{traceback.format_exc()}")
            upload_record.status = 'failed'
            upload_record.error_details = json.dumps({'error': str(e)})
            db.session.commit()
            flash('Error processing file. Please check the format and try again.', 'error')
        
        return redirect(url_for('employee_import.upload_employees'))
        
    except Exception as e:
        logger.error(f"Error in upload POST: {e}")
        flash('Server error during upload. Please try again.', 'error')
        return redirect(url_for('employee_import.upload_employees'))

# ==========================================
# TEMPLATE DOWNLOAD ROUTES - FIXED 2025-01-09
# ==========================================

@employee_import_bp.route('/download-template')
@employee_import_bp.route('/download-employee-template')
@login_required
@supervisor_required
def download_employee_template():
    """Download Excel template - qualifications are column headers, use Yes/No in cells"""
    try:
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Employee Data'
        
        # Common qualifications for your industry
        qualification_headers = [
            'Forklift Certified',
            'Crane Operator',
            'OSHA 10',
            'OSHA 30',
            'First Aid/CPR',
            'CDL License',
            'Welding Certified',
            'Electrical License',
            'Machine Operator',
            'Hazmat Certified'
        ]
        
        # Build all headers
        headers = [
            'Last Name',
            'First Name', 
            'Employee ID',
            'Crew Assigned',
            'Current Job Position',
            'Email'
        ] + qualification_headers
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers with formatting
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            # Adjust column widths
            if header == 'Email':
                ws.column_dimensions[get_column_letter(col)].width = 25
            elif header in ['Last Name', 'First Name', 'Current Job Position']:
                ws.column_dimensions[get_column_letter(col)].width = 18
            elif col > 6:  # Qualification columns
                ws.column_dimensions[get_column_letter(col)].width = 12
            else:
                ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
        
        # Add data validation for Yes/No columns
        yes_no_validation = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=True,
            showDropDown=True,
            errorTitle='Invalid Entry',
            error='Please select Yes or No'
        )
        
        # Apply validation to qualification columns (columns 7 onwards)
        for col in range(7, len(headers) + 1):
            col_letter = get_column_letter(col)
            yes_no_validation.add(f'{col_letter}2:{col_letter}1000')
        
        ws.add_data_validation(yes_no_validation)
        
        # Add sample data rows
        sample_data = [
            ['Smith', 'John', 'EMP001', 'A', 'Operator', 'john.smith@company.com', 
             'Yes', 'No', 'Yes', 'No', 'Yes', 'No', 'No', 'No', 'Yes', 'No'],
            ['Johnson', 'Sarah', 'EMP002', 'B', 'Lead Operator', 'sarah.j@company.com',
             'Yes', 'Yes', 'No', 'Yes', 'Yes', 'No', 'No', 'No', 'Yes', 'Yes'],
            ['Williams', 'Michael', 'EMP003', 'C', 'Technician', 'mike.w@company.com',
             'No', 'No', 'Yes', 'No', 'No', 'Yes', 'No', 'Yes', 'No', 'No'],
            ['Brown', 'Emily', 'EMP004', 'D', 'Supervisor', 'emily.b@company.com',
             'Yes', 'No', 'Yes', 'Yes', 'Yes', 'No', 'No', 'No', 'No', 'Yes'],
        ]
        
        # Add sample data with light gray fill
        example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.fill = example_fill
                cell.border = thin_border
                # Center align Yes/No values
                if col_num > 6:
                    cell.alignment = Alignment(horizontal="center")
        
        # Add instructions sheet
        ws2 = wb.create_sheet('Instructions')
        ws2.column_dimensions['A'].width = 100
        
        instructions = [
            ['EMPLOYEE UPLOAD TEMPLATE - HEADER-BASED QUALIFICATION SYSTEM'],
            [''],
            ['HOW TO USE THIS TEMPLATE:'],
            ['1. The column headers (row 1) define the qualifications/skills available'],
            ['2. For each employee, enter "Yes" if they have that qualification, "No" or leave blank if they don\'t'],
            ['3. You can modify the qualification headers to match your needs'],
            ['4. Add more qualification columns as needed - just add new headers'],
            [''],
            ['COLUMN DESCRIPTIONS:'],
            ['• Last Name - Employee\'s last name (REQUIRED)'],
            ['• First Name - Employee\'s first name (REQUIRED)'],
            ['• Employee ID - Unique identifier (REQUIRED)'],
            ['• Crew Assigned - Must be A, B, C, or D (REQUIRED)'],
            ['• Current Job Position - Job title (REQUIRED)'],
            ['• Email - Employee email (OPTIONAL but needed for login)'],
            ['• Qualification Columns - Enter Yes/No for each skill'],
            [''],
            ['CUSTOMIZING QUALIFICATIONS:'],
            ['1. Replace the "Add Qualification" headers with your specific needs'],
            ['2. Examples of qualifications to track:'],
            ['   - Certifications: OSHA 10, OSHA 30, First Aid, CPR, etc.'],
            ['   - Equipment: Forklift, Crane Operator, specific machines'],
            ['   - Licenses: CDL, Electrical, Plumbing'],
            ['   - Training: Leadership, Safety, Technical'],
            ['   - Languages: Spanish, Chinese, etc.'],
            ['3. You can add more qualification columns as needed'],
            ['4. Delete any unused qualification columns'],
            [''],
            ['VALID ENTRIES FOR QUALIFICATIONS:'],
            ['• Yes, Y, 1, True, X = Employee HAS this qualification'],
            ['• No, N, 0, False, blank = Employee DOES NOT have this qualification'],
            [''],
            ['DEFAULT PASSWORD: password123'],
            [''],
            ['TIPS:'],
            ['• Delete the sample rows before importing your data']
        ]
        
        for row_num, instruction in enumerate(instructions, 1):
            if instruction:
                cell = ws2.cell(row=row_num, column=1, value=instruction[0])
                if row_num == 1:
                    cell.font = Font(bold=True, size=14, color="366092")
                elif instruction[0].startswith(('COLUMN', 'CUSTOMIZING', 'VALID', 'DEFAULT', 'TIPS')):
                    cell.font = Font(bold=True, size=12)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'employee_upload_template_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Error creating employee template: {e}")
        flash('Error creating template. Please try again.', 'error')
        return redirect(url_for('employee_import.upload_employees'))

@employee_import_bp.route('/download-overtime-template')
@login_required
@supervisor_required
def download_overtime_template():
    """Download overtime template"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Overtime Data'
        
        headers = ['Employee ID', 'Week Start Date', 'Regular Hours', 'Overtime Hours', 'Total Hours', 'Notes']
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Add sample data
        sample_data = [
            ['EMP001', '2025-01-06', 40, 10, 50, 'Week 1'],
            ['EMP001', '2024-12-30', 40, 8, 48, 'Week 2'],
        ]
        
        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'overtime_template_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        logger.error(f"Error creating overtime template: {e}")
        flash('Error creating template', 'error')
        return redirect(url_for('employee_import.upload_employees'))

# ==========================================
# EXPORT ROUTES - ADDED 2025-01-09
# ==========================================

@employee_import_bp.route('/export-employees')
@login_required
@supervisor_required
def export_employees():
    """Export current employees to Excel"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Employee Data'
        
        # Get all active employees
        employees = Employee.query.filter_by(is_active=True).all()
        
        # Get all skills
        all_skills = Skill.query.all()
        skill_headers = [skill.name for skill in all_skills]
        
        # Create headers
        headers = ['Last Name', 'First Name', 'Employee ID', 'Crew Assigned', 'Current Job Position', 'Email'] + skill_headers
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Write employee data
        for row_num, employee in enumerate(employees, 2):
            # Split name
            name_parts = employee.name.split(' ', 1) if employee.name else ['', '']
            first_name = name_parts[0] if len(name_parts) > 0 else ''
            last_name = name_parts[1] if len(name_parts) > 1 else ''
            
            ws.cell(row=row_num, column=1, value=last_name)
            ws.cell(row=row_num, column=2, value=first_name)
            ws.cell(row=row_num, column=3, value=employee.employee_id)
            ws.cell(row=row_num, column=4, value=employee.crew)
            ws.cell(row=row_num, column=5, value=employee.position.name if employee.position else '')
            ws.cell(row=row_num, column=6, value=employee.email)
            
            # Add skills
            employee_skills = {es.skill.name for es in employee.employee_skills}
            for col, skill_name in enumerate(skill_headers, 7):
                ws.cell(row=row_num, column=col, value='Yes' if skill_name in employee_skills else 'No')
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'employees_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Error exporting employees: {e}")
        flash('Error exporting employees', 'error')
        return redirect(url_for('employee_import.upload_employees'))

@employee_import_bp.route('/export-overtime')
@login_required
@supervisor_required
def export_overtime():
    """Export overtime data to Excel"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Overtime Data'
        
        # Get all overtime records
        overtime_records = OvertimeHistory.query.join(Employee).order_by(
            Employee.employee_id, OvertimeHistory.week_start_date
        ).all()
        
        # Headers
        headers = ['Employee ID', 'Employee Name', 'Week Start Date', 'Regular Hours', 'Overtime Hours', 'Total Hours']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Write data
        for row_num, ot in enumerate(overtime_records, 2):
            ws.cell(row=row_num, column=1, value=ot.employee.employee_id)
            ws.cell(row=row_num, column=2, value=ot.employee.name)
            ws.cell(row=row_num, column=3, value=ot.week_start_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=4, value=ot.regular_hours)
            ws.cell(row=row_num, column=5, value=ot.overtime_hours)
            ws.cell(row=row_num, column=6, value=ot.total_hours)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'overtime_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Error exporting overtime: {e}")
        flash('Error exporting overtime data', 'error')
        return redirect(url_for('employee_import.upload_employees'))

# ==========================================
# AJAX ENDPOINTS FOR UPLOAD HISTORY - ADDED 2025-01-09
# ==========================================

@employee_import_bp.route('/upload-details/<int:upload_id>')
@login_required
@supervisor_required
def get_upload_details(upload_id):
    """Get upload details for modal"""
    try:
        upload = FileUpload.query.get_or_404(upload_id)
        
        # Parse error details if JSON
        error_details = None
        if upload.error_details:
            try:
                error_details = json.loads(upload.error_details)
            except:
                error_details = upload.error_details
        
        return jsonify({
            'id': upload.id,
            'filename': upload.filename,
            'upload_type': upload.upload_type or 'employee',
            'status': upload.status or 'completed',
            'file_size': upload.file_size,
            'records_processed': upload.records_processed or 0,
            'successful_records': upload.successful_records or 0,
            'created_at': upload.uploaded_at.isoformat() if upload.uploaded_at else None,
            'uploaded_by': upload.uploaded_by.name if upload.uploaded_by else 'Unknown',
            'error_log': error_details,
            'original_file_path': bool(upload.file_path and os.path.exists(upload.file_path))
        })
        
    except Exception as e:
        logger.error(f"Error getting upload details: {e}")
        return jsonify({'error': str(e)}), 500

@employee_import_bp.route('/upload-errors/<int:upload_id>')
@login_required
@supervisor_required
def get_upload_errors(upload_id):
    """Get upload error details"""
    try:
        upload = FileUpload.query.get_or_404(upload_id)
        
        errors = []
        if upload.error_details:
            try:
                error_data = json.loads(upload.error_details)
                if isinstance(error_data, dict):
                    errors = error_data.get('errors', [])
                elif isinstance(error_data, list):
                    errors = error_data
            except:
                errors = [{'message': upload.error_details}]
        
        # Format errors for display
        formatted_errors = []
        for error in errors:
            if isinstance(error, str):
                formatted_errors.append({'message': error, 'type': 'Error'})
            else:
                formatted_errors.append(error)
        
        return jsonify({
            'errors': formatted_errors,
            'error_count': len(formatted_errors)
        })
        
    except Exception as e:
        logger.error(f"Error getting upload errors: {e}")
        return jsonify({'error': str(e)}), 500

@employee_import_bp.route('/download-upload/<int:upload_id>')
@login_required
@supervisor_required
def download_upload(upload_id):
    """Download original uploaded file"""
    try:
        upload = FileUpload.query.get_or_404(upload_id)
        
        if upload.file_path and os.path.exists(upload.file_path):
            return send_file(
                upload.file_path,
                as_attachment=True,
                download_name=upload.filename
            )
        else:
            flash('Original file not found', 'error')
            return redirect(url_for('employee_import.upload_history'))
            
    except Exception as e:
        logger.error(f"Error downloading upload: {e}")
        flash('Error downloading file', 'error')
        return redirect(url_for('employee_import.upload_history'))

@employee_import_bp.route('/delete-upload/<int:upload_id>', methods=['DELETE'])
@login_required
@supervisor_required
def delete_upload(upload_id):
    """Delete upload record"""
    try:
        upload = FileUpload.query.get_or_404(upload_id)
        
        # Delete file if exists
        if upload.file_path and os.path.exists(upload.file_path):
            try:
                os.remove(upload.file_path)
            except:
                pass
        
        db.session.delete(upload)
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        logger.error(f"Error deleting upload: {e}")
        return jsonify({'error': str(e)}), 500

# ==========================================
# UPLOAD HISTORY ROUTE - FIXED
# ==========================================

@employee_import_bp.route('/upload-history')
@login_required
@supervisor_required
def upload_history():
    """View upload history - FIXED to handle empty results"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = 20
        
        # Build query with filters
        query = FileUpload.query
        
        # Apply filters from request args
        search = request.args.get('search', '')
        if search:
            query = query.filter(
                or_(
                    FileUpload.filename.contains(search)
                )
            )
        
        upload_type = request.args.get('upload_type', '')
        if upload_type:
            query = query.filter_by(upload_type=upload_type)
        
        status = request.args.get('status', '')
        if status:
            query = query.filter_by(status=status)
        
        # Order and paginate
        uploads = query.order_by(FileUpload.uploaded_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        # Count statistics
        successful_uploads = FileUpload.query.filter_by(status='completed').count()
        failed_uploads = FileUpload.query.filter_by(status='failed').count()
        
        # Check if template exists
        try:
            return render_template(
                'upload_history.html',
                uploads=uploads,
                successful_uploads=successful_uploads,
                failed_uploads=failed_uploads
            )
        except:
            # If template doesn't exist, show simple list
            return render_simple_history_page(uploads, successful_uploads, failed_uploads)
        
    except Exception as e:
        logger.error(f"Error in upload_history: {e}")
        # Don't redirect to upload_employees, show error
        return render_simple_history_page(None, 0, 0, error=str(e))

def render_simple_history_page(uploads, successful, failed, error=None):
    """Render a simple history page if template is missing"""
    if error:
        content = f'<div class="alert alert-danger">Error loading history: {error}</div>'
    elif not uploads or uploads.total == 0:
        content = '''
        <div class="alert alert-info">
            <h4>No Upload History</h4>
            <p>No files have been uploaded yet.</p>
            <a href="/upload-employees" class="btn btn-primary">Upload First File</a>
        </div>
        '''
    else:
        rows = []
        for upload in uploads.items:
            rows.append(f'''
            <tr>
                <td>{upload.filename}</td>
                <td>{upload.upload_type or 'employee'}</td>
                <td>{upload.status or 'unknown'}</td>
                <td>{upload.records_processed or 0}</td>
                <td>{upload.uploaded_at.strftime('%Y-%m-%d %H:%M') if upload.uploaded_at else 'N/A'}</td>
            </tr>
            ''')
        
        content = f'''
        <table class="table">
            <thead>
                <tr>
                    <th>Filename</th>
                    <th>Type</th>
                    <th>Status</th>
                    <th>Records</th>
                    <th>Uploaded</th>
                </tr>
            </thead>
            <tbody>
                {''.join(rows)}
            </tbody>
        </table>
        '''
    
    html = f'''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Upload History</title>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
    </head>
    <body>
        <nav class="navbar navbar-dark bg-primary">
            <div class="container-fluid">
                <span class="navbar-brand">Upload History</span>
                <div>
                    <a href="/upload-employees" class="btn btn-light btn-sm">Back to Upload</a>
                    <a href="/dashboard" class="btn btn-light btn-sm">Dashboard</a>
                </div>
            </div>
        </nav>
        
        <div class="container mt-4">
            <h2>Upload History</h2>
            <div class="row mb-3">
                <div class="col-md-4">
                    <div class="card text-center">
                        <div class="card-body">
                            <h3>{successful}</h3>
                            <p>Successful</p>
                        </div>
                    </div>
                </div>
                <div class="col-md-4">
                    <div class="card text-center">
                        <div class="card-body">
                            <h3>{failed}</h3>
                            <p>Failed</p>
                        </div>
                    </div>
                </div>
                <div class="col-md-4">
                    <div class="card text-center">
                        <div class="card-body">
                            <h3>{successful + failed}</h3>
                            <p>Total</p>
                        </div>
                    </div>
                </div>
            </div>
            {content}
        </div>
    </body>
    </html>
    '''
    return make_response(html)

# ==========================================
# HELPER FUNCTION FOR SIMPLE UPLOAD PAGE
# ==========================================

def render_simple_upload_page(stats, recent_uploads):
    """Render a simple upload page inline if templates are missing"""
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Upload Employees</title>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
        <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.8.1/font/bootstrap-icons.css">
    </head>
    <body>
        <nav class="navbar navbar-dark bg-primary">
            <div class="container-fluid">
                <span class="navbar-brand mb-0 h1">Workforce Scheduler</span>
                <a href="/dashboard" class="btn btn-light btn-sm">Back to Dashboard</a>
            </div>
        </nav>
        
        <div class="container mt-5">
            <h2><i class="bi bi-upload"></i> Employee Data Upload</h2>
            
            <div class="alert alert-info">
                <strong>New System:</strong> Column headers define qualifications. Just enter Yes/No for each employee.
            </div>
            
            <div class="row mt-4">
                <div class="col-md-8">
                    <div class="card">
                        <div class="card-header bg-primary text-white">
                            <h5 class="mb-0">Upload Excel File</h5>
                        </div>
                        <div class="card-body">
                            <form method="POST" action="/upload-employees" enctype="multipart/form-data">
                                <div class="mb-3">
                                    <label for="file" class="form-label">Select Excel File</label>
                                    <input type="file" class="form-control" id="file" name="file" accept=".xlsx,.xls" required>
                                    <div class="form-text">Download template to see the header-based format</div>
                                </div>
                                
                                <div class="mb-3">
                                    <div class="form-check">
                                        <input class="form-check-input" type="checkbox" id="replaceAll" name="replaceAll">
                                        <label class="form-check-label" for="replaceAll">
                                            Replace all existing data (use with caution)
                                        </label>
                                    </div>
                                </div>
                                
                                <button type="submit" class="btn btn-primary">
                                    <i class="bi bi-upload"></i> Upload File
                                </button>
                                <a href="/download-employee-template" class="btn btn-secondary">
                                    <i class="bi bi-download"></i> Download Template
                                </a>
                                <a href="/upload-history" class="btn btn-info">
                                    <i class="bi bi-clock-history"></i> History
                                </a>
                            </form>
                        </div>
                    </div>
                </div>
                
                <div class="col-md-4">
                    <div class="card">
                        <div class="card-header bg-info text-white">
                            <h5 class="mb-0">Statistics</h5>
                        </div>
                        <div class="card-body">
                            <p>Total Employees: <strong>{stats.get('total_employees', 0)}</strong></p>
                            <p>Total Skills: <strong>{stats.get('total_skills', 0)}</strong></p>
                            <p>Employees with Skills: <strong>{stats.get('employees_with_skills', 0)}</strong></p>
                            <hr>
                            <small class="text-muted">Skills are used for overtime eligibility</small>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </body>
    </html>
    """
    return make_response(html)

# ==========================================
# TEST ROUTE
# ==========================================

@employee_import_bp.route('/test-upload-route')
def test_upload_route():
    """Test route to verify blueprint is loaded"""
    return jsonify({
        'status': 'Employee import blueprint is working!',
        'routes_available': [
            '/upload-employees',
            '/validate-upload',
            '/download-employee-template',
            '/download-template',
            '/upload-history',
            '/export-employees',
            '/export-overtime',
            '/upload-details/<id>',
            '/upload-errors/<id>',
            '/download-upload/<id>',
            '/delete-upload/<id>'
        ],
        'system': 'Header-based qualifications - column headers define skills, cells contain Yes/No',
        'default_password': 'password123',
        'authenticated': current_user.is_authenticated,
        'is_supervisor': current_user.is_supervisor if current_user.is_authenticated else False,
        'upload_folder': current_app.config.get('UPLOAD_FOLDER'),
        'timestamp': datetime.now().isoformat()
    })

# Log successful blueprint loading
logger.info("Employee import blueprint loaded - Header-based qualification system active - All routes added 2025-09-08")
